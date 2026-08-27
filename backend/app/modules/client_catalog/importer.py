"""Import CIMMYT client-controlled catalogs from Excel."""

import io
import logging
from typing import Any, Dict, List

import openpyxl

from app.core.database import transaction

logger = logging.getLogger(__name__)

CATALOG_SHEET = "04_Value_Catalogs"
VALUES_SHEET = "05_Catalog_Values"


class CatalogImportError(ValueError):
    """The workbook cannot be interpreted as a catalog workbook."""


def _text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() in ("nan", "none", "-"):
        return ""

    return text


def _order(value: Any) -> int:
    try:
        return int(float(_text(value)))
    except (TypeError, ValueError):
        return 10000


def _read_sheet(ws) -> List[Dict[str, Any]]:
    """Read a sheet whose real header may be below title/description rows."""

    rows = list(ws.iter_rows(values_only=True))

    header_index = None

    # CIMMYT sheets normally have title/description before row 4.
    # Detect the header instead of assuming row 4.
    for index, row in enumerate(rows[:15]):

        filled = [_text(cell) for cell in row if _text(cell)]

        if len(filled) >= 2:
            header_index = index
            break

    if header_index is None:
        return []

    headers = [_text(cell) for cell in rows[header_index]]

    result = []

    for row in rows[header_index + 1:]:

        record = {}

        for header, value in zip(headers, row):

            if header:
                record[header] = _text(value)

        if any(record.values()):
            result.append(record)

    return result


def read_catalog_workbook(data: bytes) -> Dict[str, List[Dict[str, Any]]]:

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise CatalogImportError(
            f"Could not open Excel workbook: {exc}"
        ) from exc

    sheets = set(workbook.sheetnames)

    missing = []

    if CATALOG_SHEET not in sheets:
        missing.append(CATALOG_SHEET)

    if VALUES_SHEET not in sheets:
        missing.append(VALUES_SHEET)

    if missing:
        raise CatalogImportError(
            "This workbook is missing required CIMMYT sheet(s): "
            + ", ".join(missing)
            + ". Found: "
            + ", ".join(workbook.sheetnames)
        )

    catalogs = _read_sheet(workbook[CATALOG_SHEET])
    values = _read_sheet(workbook[VALUES_SHEET])

    if not catalogs:
        raise CatalogImportError(
            f"Sheet {CATALOG_SHEET} exists but contains no catalog rows."
        )

    if not values:
        logger.warning(
            "Sheet %s contains no catalog values.",
            VALUES_SHEET,
        )

    return {
        "catalogs": catalogs,
        "values": values,
    }


def import_catalog_workbook(
    data: bytes,
    source: str = "",
) -> Dict[str, int]:

    sheets = read_catalog_workbook(data)

    catalogs = sheets["catalogs"]
    values = sheets["values"]

    catalogs_added = 0
    catalogs_updated = 0
    values_added = 0
    values_updated = 0
    values_skipped = 0

    with transaction() as cur:

        # ---------------------------------------------------------------
        # Catalog definitions
        # ---------------------------------------------------------------

        for row in catalogs:

            catalog_id = _text(row.get("Catalog ID"))

            if not catalog_id:
                continue

            name = (
                _text(row.get("Catalog Name"))
                or catalog_id
            )

            description = _text(
                row.get("Description")
                or row.get("Definition")
            )

            version = _text(
                row.get("Version")
            )

            status = _text(
                row.get("Status")
            )

            cur.execute(
                """
                SELECT 1
                FROM client_catalog
                WHERE catalog_id = %s
                """,
                (catalog_id,),
            )

            exists = cur.fetchone() is not None

            cur.execute(
                """
                INSERT INTO client_catalog
                    (
                        catalog_id,
                        name,
                        description,
                        version,
                        status,
                        source,
                        imported_on,
                        updated_on
                    )
                VALUES
                    (%s, %s, %s, %s, %s, %s,
                     CURRENT_TIMESTAMP,
                     CURRENT_TIMESTAMP)

                ON CONFLICT (catalog_id)
                DO UPDATE SET
                    name = EXCLUDED.name,
                    description = EXCLUDED.description,
                    version = EXCLUDED.version,
                    status = EXCLUDED.status,
                    source = EXCLUDED.source,
                    updated_on = CURRENT_TIMESTAMP
                """,
                (
                    catalog_id,
                    name,
                    description,
                    version,
                    status,
                    source,
                ),
            )

            if exists:
                catalogs_updated += 1
            else:
                catalogs_added += 1

        # ---------------------------------------------------------------
        # Catalog values
        # ---------------------------------------------------------------

        known_catalogs = set()

        cur.execute(
            "SELECT catalog_id FROM client_catalog"
        )

        for row in cur.fetchall():
            known_catalogs.add(row["catalog_id"])

        for row in values:

            catalog_id = _text(
                row.get("Catalog ID")
            )

            code = _text(
                row.get("Code")
            )

            if not catalog_id or not code:
                values_skipped += 1
                continue

            if catalog_id not in known_catalogs:
                logger.warning(
                    "Skipping value %s: catalog %s does not exist",
                    code,
                    catalog_id,
                )
                values_skipped += 1
                continue

            label = (
                _text(row.get("Preferred Label EN"))
                or _text(row.get("Preferred Label"))
                or code
            )

            definition = _text(
                row.get("Definition")
            )

            parent_code = _text(
                row.get("Parent Code")
            ) or None

            display_order = _order(
                row.get("Display Order")
            )

            status = _text(
                row.get("Status")
            )

            valid_from = _text(
                row.get("Valid From")
            )

            valid_to = _text(
                row.get("Valid To")
            )

            cur.execute(
                """
                SELECT 1
                FROM client_catalog_value
                WHERE catalog_id = %s
                  AND code = %s
                """,
                (
                    catalog_id,
                    code,
                ),
            )

            exists = cur.fetchone() is not None

            cur.execute(
                """
                INSERT INTO client_catalog_value
                    (
                        catalog_id,
                        code,
                        label,
                        definition,
                        parent_code,
                        display_order,
                        status,
                        valid_from,
                        valid_to,
                        imported_on,
                        updated_on
                    )
                VALUES
                    (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        CURRENT_TIMESTAMP,
                        CURRENT_TIMESTAMP
                    )

                ON CONFLICT (catalog_id, code)
                DO UPDATE SET
                    label = EXCLUDED.label,
                    definition = EXCLUDED.definition,
                    parent_code = EXCLUDED.parent_code,
                    display_order = EXCLUDED.display_order,
                    status = EXCLUDED.status,
                    valid_from = EXCLUDED.valid_from,
                    valid_to = EXCLUDED.valid_to,
                    updated_on = CURRENT_TIMESTAMP
                """,
                (
                    catalog_id,
                    code,
                    label,
                    definition,
                    parent_code,
                    display_order,
                    status,
                    valid_from,
                    valid_to,
                ),
            )

            if exists:
                values_updated += 1
            else:
                values_added += 1

    return {
        "catalogs_total": len(catalogs),
        "values_total": len(values),
        "catalogs_added": catalogs_added,
        "catalogs_updated": catalogs_updated,
        "values_added": values_added,
        "values_updated": values_updated,
        "values_skipped": values_skipped,
    }


def list_catalogs() -> List[Dict[str, Any]]:

    with transaction() as cur:

        cur.execute(
            """
            SELECT
                catalog_id,
                name,
                description,
                version,
                status,
                source,
                imported_on,
                updated_on
            FROM client_catalog
            ORDER BY name, catalog_id
            """
        )

        return [dict(row) for row in cur.fetchall()]


def get_catalog(catalog_id: str) -> Dict[str, Any] | None:

    with transaction() as cur:

        cur.execute(
            """
            SELECT
                catalog_id,
                name,
                description,
                version,
                status,
                source,
                imported_on,
                updated_on
            FROM client_catalog
            WHERE catalog_id = %s
            """,
            (catalog_id,),
        )

        row = cur.fetchone()

        return dict(row) if row else None


def get_values(
    catalog_id: str,
    parent_code: str | None = None,
) -> List[Dict[str, Any]]:

    with transaction() as cur:

        if parent_code is None:

            cur.execute(
                """
                SELECT
                    code,
                    label,
                    definition,
                    parent_code,
                    display_order,
                    status
                FROM client_catalog_value
                WHERE catalog_id = %s
                ORDER BY display_order, code
                """,
                (catalog_id,),
            )

        else:

            cur.execute(
                """
                SELECT
                    code,
                    label,
                    definition,
                    parent_code,
                    display_order,
                    status
                FROM client_catalog_value
                WHERE catalog_id = %s
                  AND parent_code = %s
                ORDER BY display_order, code
                """,
                (
                    catalog_id,
                    parent_code,
                ),
            )

        return [dict(row) for row in cur.fetchall()]