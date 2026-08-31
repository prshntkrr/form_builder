"""Import the client's own catalogue workbook — the "Catalogs" sheet format.

Deliberately separate from importer.py, which reads the CIMMYT Controlled
Vocabulary workbook (04_Value_Catalogs / 05_Catalog_Values). Two workbooks, two
readers, one set of tables: whichever produced a catalogue, it behaves the same
afterwards.

The sheet is one row per value, grouped by list:

    List              Variable        Label Spanish     Label ENG
    SiNo_list         Si              Si                Yes
    SiNo_list         No              No                No
    Escolaridad_list  Sin_estudios    Sin estudios      Without studies

    List           the catalogue's id
    Variable       the code — what an answer stores, and the one stable thing
    Label Spanish  what a Spanish reader sees
    Label ENG      what an English reader sees

One row per value, not one per language. `Si` is one code with two labels; it
would be a different answer if it were two rows, and the same person answering
in two languages must produce the same data.

Nothing here translates. A label that the workbook leaves empty stays empty —
the other language is shown in its place at render time, which is a fallback,
not a translation.
"""
import io
import logging
import warnings
from typing import Any, Dict, List, Optional

import openpyxl
from psycopg2.extras import Json

from app.core.database import transaction

logger = logging.getLogger(__name__)

SHEET = "Catalogs"

LIST_COLUMN = "list"
CODE_COLUMN = "variable"

# The label columns, and the language each one is written in. The workbook's own
# headings are the evidence; nothing is assumed about which languages a client
# supplies, and a heading we do not recognise is simply not a label column.
LABEL_COLUMNS = {
    "label spanish": "es",
    "label español": "es",
    "label esp": "es",
    "etiqueta": "es",
    "label eng": "en",
    "label english": "en",
    "etiqueta eng": "en",
    "label fr": "fr",
    "label français": "fr",
    "label pt": "pt",
}

# Which language's label becomes the value's own `label` — the one shown when
# no language is asked for, and the one older code paths read. First match wins.
PRIMARY_ORDER = ("es", "en", "fr", "pt")


class EagrologyCatalogError(ValueError):
    """The workbook is not a supported e-Agrology catalogue workbook."""


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none", "-") else text


def _headings(row) -> Dict[str, int]:
    """Which column holds what, by heading, case and spacing ignored."""
    return {_text(cell).lower(): index for index, cell in enumerate(row) if _text(cell)}


def _catalogs_sheet(workbook):
    for name in workbook.sheetnames:
        if name.strip().lower() == SHEET.lower():
            return workbook[name]
    return None


def is_eagrology_workbook(data: bytes) -> bool:
    """Whether this is the client's own catalogue workbook.

    Asked before the CIMMYT reader is, so a workbook in this shape is never
    handed to a reader that would demand sheets it does not have.
    """
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

            sheet = _catalogs_sheet(workbook)
            if sheet is None:
                return False

            first = next(sheet.iter_rows(max_row=1, values_only=True), ())
            headings = _headings(first)

            # A list column, a code column, and at least one label column.
            return (
                LIST_COLUMN in headings
                and CODE_COLUMN in headings
                and any(heading in LABEL_COLUMNS for heading in headings)
            )
    except Exception:
        return False


def read_workbook(data: bytes) -> Dict[str, Any]:
    """The workbook as catalogues and their values. Reads only; writes nothing."""
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            sheet = _catalogs_sheet(workbook)

            if sheet is None:
                raise EagrologyCatalogError(
                    f"This workbook has no '{SHEET}' sheet. Found: "
                    + ", ".join(workbook.sheetnames)
                )

            rows = list(sheet.iter_rows(values_only=True))
    except EagrologyCatalogError:
        raise
    except Exception as exc:
        raise EagrologyCatalogError(f"Could not read the workbook: {exc}") from exc

    if not rows:
        raise EagrologyCatalogError(f"The '{SHEET}' sheet is empty.")

    headings = _headings(rows[0])

    if LIST_COLUMN not in headings or CODE_COLUMN not in headings:
        raise EagrologyCatalogError(
            f"The '{SHEET}' sheet needs a List column and a Variable column. Found: "
            + ", ".join(_text(cell) for cell in rows[0] if _text(cell))
        )

    label_columns = {
        language: headings[heading]
        for heading, language in LABEL_COLUMNS.items()
        if heading in headings
    }

    if not label_columns:
        raise EagrologyCatalogError(
            f"The '{SHEET}' sheet has no label column, so its values would have "
            f"nothing to show. Expected one of: Label Spanish, Label ENG."
        )

    catalogs: Dict[str, Dict[str, Any]] = {}
    seen: set = set()
    duplicates: List[str] = []
    headers_skipped = 0

    def cell(row, index: Optional[int]) -> str:
        return _text(row[index]) if index is not None and index < len(row) else ""

    for row in rows[1:]:
        catalog_id = cell(row, headings[LIST_COLUMN])
        if not catalog_id:
            continue

        catalog = catalogs.setdefault(catalog_id, {"catalog_id": catalog_id, "values": []})

        code = cell(row, headings[CODE_COLUMN])
        if not code:
            # A row naming the list and nothing else — the group heading above
            # its values. It creates the catalogue, not a value in it.
            headers_skipped += 1
            continue

        if (catalog_id, code) in seen:
            # The workbook says the same code twice in one list. Keeping the
            # first is the only safe reading: a code identifies one value.
            duplicates.append(f"{catalog_id}/{code}")
            continue
        seen.add((catalog_id, code))

        labels = {
            language: cell(row, index)
            for language, index in label_columns.items()
            if cell(row, index)
        }

        catalog["values"].append({
            "code": code,
            "labels": labels,
            "display_order": len(catalog["values"]) + 1,
        })

    if not catalogs:
        raise EagrologyCatalogError(f"The '{SHEET}' sheet holds no catalogue rows.")

    return {
        "catalogs": list(catalogs.values()),
        "languages": sorted(label_columns),
        "duplicates": duplicates,
        "headers_skipped": headers_skipped,
    }


def _primary(labels: Dict[str, str], code: str) -> str:
    """The label shown when no language is asked for."""
    for language in PRIMARY_ORDER:
        if labels.get(language):
            return labels[language]
    for label in labels.values():
        if label:
            return label
    return code


def import_workbook(data: bytes, source: str = "") -> Dict[str, Any]:
    """Load the workbook into the catalogue tables.

    Idempotent: catalogues and values are matched on the client's own ids, so
    re-importing updates and adds rather than duplicating.

    An **Approved** catalogue is treated as settled. New values may be added to
    it and a language it was missing may be filled in — neither changes what an
    existing code means. Rewording a code it already has is refused and reported
    as a conflict instead, because answers already carry that code.
    """
    read = read_workbook(data)

    catalogs_added = catalogs_updated = 0
    values_added = values_updated = values_skipped = 0
    conflicts: List[Dict[str, str]] = []

    values_total = sum(len(c["values"]) for c in read["catalogs"])

    with transaction() as cur:
        for catalog in read["catalogs"]:
            catalog_id = catalog["catalog_id"]

            cur.execute(
                "SELECT status FROM client_catalog WHERE catalog_id = %s",
                (catalog_id,),
            )
            existing = cur.fetchone()

            if existing is None:
                # A newly imported catalogue is a Candidate: it exists and can be
                # used, but nobody has yet said it is settled.
                cur.execute(
                    """
                    INSERT INTO client_catalog
                        (catalog_id, name, description, version, status, source,
                         created_by, imported_on, updated_on)
                    VALUES (%s, %s, '', '1.0', 'Candidate', %s, '',
                            CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """,
                    (catalog_id, catalog_id, source),
                )
                catalogs_added += 1
                approved = False
            else:
                # Its name, version and status are the client's to set in the
                # builder. A re-import brings values, not a change of standing.
                cur.execute(
                    "UPDATE client_catalog SET source = %s, updated_on = CURRENT_TIMESTAMP "
                    "WHERE catalog_id = %s",
                    (source, catalog_id),
                )
                catalogs_updated += 1
                approved = _text(existing["status"]).lower() == "approved"

            for value in catalog["values"]:
                code = value["code"]
                labels = value["labels"]
                label = _primary(labels, code)

                cur.execute(
                    "SELECT label, labels FROM client_catalog_value "
                    "WHERE catalog_id = %s AND code = %s",
                    (catalog_id, code),
                )
                current = cur.fetchone()

                if current is None:
                    cur.execute(
                        """
                        INSERT INTO client_catalog_value
                            (catalog_id, code, label, labels, definition,
                             display_order, status, imported_on, updated_on)
                        VALUES (%s, %s, %s, %s, '', %s, 'Active',
                                CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                        """,
                        (catalog_id, code, label, Json(labels), value["display_order"]),
                    )
                    values_added += 1
                    continue

                held = dict(current["labels"] or {})

                if approved:
                    # Filling in a language the value did not have is not a
                    # change of meaning, so it is allowed. Rewording one it
                    # already has is, so it is not.
                    reworded = [
                        language for language, text in labels.items()
                        if held.get(language) and held[language] != text
                    ]
                    if reworded or (current["label"] and current["label"] != label
                                    and not held):
                        conflicts.append({
                            "catalog_id": catalog_id,
                            "code": code,
                            "held": held.get(reworded[0]) if reworded else current["label"],
                            "workbook": labels.get(reworded[0]) if reworded else label,
                            "reason": f"{catalog_id} is Approved and '{code}' already means "
                                      f"something here. The workbook was not applied to it.",
                        })
                        values_skipped += 1
                        continue

                    added = {k: v for k, v in labels.items() if not held.get(k)}
                    if not added:
                        values_skipped += 1
                        continue

                    held.update(added)
                    cur.execute(
                        "UPDATE client_catalog_value SET labels = %s, "
                        "updated_on = CURRENT_TIMESTAMP WHERE catalog_id = %s AND code = %s",
                        (Json(held), catalog_id, code),
                    )
                    values_updated += 1
                    continue

                held.update(labels)
                cur.execute(
                    """
                    UPDATE client_catalog_value
                    SET label = %s, labels = %s, display_order = %s,
                        updated_on = CURRENT_TIMESTAMP
                    WHERE catalog_id = %s AND code = %s
                    """,
                    (label, Json(held), value["display_order"], catalog_id, code),
                )
                values_updated += 1

    logger.info(
        "Imported %s catalogue(s) and %s value(s) from %s",
        catalogs_added + catalogs_updated, values_added + values_updated, source or "a workbook",
    )

    return {
        "format": "eagrology",
        "languages": read["languages"],
        "catalogs_total": len(read["catalogs"]),
        "catalogs_added": catalogs_added,
        "catalogs_updated": catalogs_updated,
        "values_total": values_total,
        "values_added": values_added,
        "values_updated": values_updated,
        "values_skipped": values_skipped,
        # The workbook's own group rows: a list named with no value beside it.
        # Reported so the counts add up rather than looking as if rows vanished.
        "headers_skipped": read["headers_skipped"],
        "duplicates": read["duplicates"][:50],
        "duplicate_count": len(read["duplicates"]),
        "conflicts": conflicts[:50],
        "conflict_count": len(conflicts),
    }
