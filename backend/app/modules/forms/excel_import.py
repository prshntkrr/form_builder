"""Reading a CIMMYT Controlled Vocabulary workbook into Form JSON.

The workbook is an existing form definition written by the client, so it is the
authority on everything it states: the fields, their types, the wording, the
controlled lists and the language they are written in. Nothing here invents any
of that, and nothing here translates.

The sheets that carry a form, and what each contributes:

    14_Profiles             one profile is one form — which variables, in what
                            order, and whether each is required
    03_Variables            the field itself: name, definition, data type, the
                            unit or catalog it uses
    11_Question_Items       the wording a person actually reads, and the
                            language tag it is written in
    05_Catalog_Values       the permitted values of a controlled list, with
                            `Parent Code` for a list nested under another
    04_Value_Catalogs       what each catalog is called
    06_Units                the unit a numeric variable is measured in
    10_Multilingual_Labels  translations the client already has
    09_External_Mappings    standards the client has already mapped

Header rows are found rather than assumed: each sheet opens with a title and a
description before the real header, and a template may gain rows above it.
"""
import logging
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# The sheets this reader uses. A workbook missing an optional one still imports.
PROFILES = "14_Profiles"
VARIABLES = "03_Variables"
QUESTIONS = "11_Question_Items"
CATALOGS = "04_Value_Catalogs"
CATALOG_VALUES = "05_Catalog_Values"
UNITS = "06_Units"
LABELS = "10_Multilingual_Labels"
MAPPINGS = "09_External_Mappings"

REQUIRED_SHEETS = (VARIABLES,)

# The workbook's data types, as its own `03_Variables.Data Type` column uses
# them, mapped onto the field types this application stores. Anything the
# workbook says that is not listed becomes text — the safest thing to store an
# unknown value in, and never a guess about what it meant.
DATA_TYPES = {
    "decimal": "decimal",
    "float": "decimal",
    "double": "decimal",
    "real": "decimal",
    "integer": "number",
    "int": "number",
    "count": "number",
    "boolean": "boolean",
    "bool": "boolean",
    "code": "select",
    "categorical": "select",
    "coded": "select",
    "text": "text",
    "string": "text",
    "memo": "textarea",
    "date": "date",
    "datetime": "datetime",
    "time": "time",
}

# What `14_Profiles.Requirement Level` has to say for a field to be required.
REQUIRED_LEVELS = {"required", "mandatory", "must"}


class WorkbookProblem(ValueError):
    """The workbook could not be read as a form definition."""


def _text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in ("nan", "none", "-") else text


def _slug(text: str) -> str:
    from app.modules.forms.form_schema import slugify_identifier

    return slugify_identifier(text, "")


def _sheet_rows(worksheet) -> List[Dict[str, Any]]:
    """One dict per data row, keyed by the sheet's own column headings.

    The header is the first row carrying several values: every sheet opens with
    a title and a description, and neither is a header.
    """
    rows = list(worksheet.iter_rows(values_only=True))

    header_at = None
    for index, row in enumerate(rows[:12]):
        filled = [cell for cell in row if _text(cell)]
        # Two or more headings. The title and description rows above hold one
        # cell each, and a narrow sheet — Units has three columns — is still a
        # header and must not be skipped for being small.
        if len(filled) >= 2:
            header_at = index
            break

    if header_at is None:
        return []

    headings = [_text(cell) for cell in rows[header_at]]
    table = []
    for row in rows[header_at + 1:]:
        record = {}
        for heading, cell in zip(headings, row):
            if heading and _text(cell):
                record[heading] = _text(cell)
        if record:
            table.append(record)
    return table


def read_workbook(data: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """Every sheet of the workbook, as rows keyed by its own column headings."""
    try:
        import io
        import warnings

        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise WorkbookProblem(
            "openpyxl is not installed. Add it with: pip install openpyxl"
        ) from exc

    try:
        with warnings.catch_warnings():
            # The template carries data-validation rules openpyxl cannot model.
            # They do not affect the values, which is all this reads.
            warnings.simplefilter("ignore")
            workbook = openpyxl.load_workbook(
                io.BytesIO(data), read_only=True, data_only=True
            )
    except Exception as exc:
        raise WorkbookProblem(f"That file could not be opened as .xlsx: {exc}") from exc

    with warnings.catch_warnings():
        # read_only defers the parse until the rows are walked, so the
        # data-validation warning surfaces here rather than at load.
        warnings.simplefilter("ignore")
        sheets = {name: _sheet_rows(workbook[name]) for name in workbook.sheetnames}

    missing = [name for name in REQUIRED_SHEETS if not sheets.get(name)]
    if missing:
        raise WorkbookProblem(
            f"The workbook has no {', '.join(missing)} sheet with any rows. "
            "This reader expects a CIMMYT Controlled Vocabulary workbook."
        )
    return sheets


# --------------------------------------------------------------------------- #
# the pieces a form is built from
# --------------------------------------------------------------------------- #
def _catalog_options(values: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    """The permitted values of each catalog, in the order the workbook gives.

    `Parent Code` is carried through untouched. It is how the client expresses a
    list nested under another — a district under its region — and dropping it
    would lose a relationship this reader has no business re-deriving.
    """
    by_catalog: Dict[str, List[Dict[str, str]]] = {}

    for row in values:
        catalog_id = _text(row.get("Catalog ID"))
        code = _text(row.get("Code"))
        if not catalog_id or not code:
            continue
        if _text(row.get("Status")).lower() in ("deprecated", "retired", "withdrawn"):
            continue

        option = {
            "label": _text(row.get("Preferred Label EN")) or code,
            "value": code,
        }
        parent = _text(row.get("Parent Code"))
        if parent:
            option["parent_code"] = parent
        definition = _text(row.get("Definition"))
        if definition:
            option["description"] = definition

        by_catalog.setdefault(catalog_id, []).append((
            _order(row.get("Display Order")), option,
        ))

    ordered: Dict[str, List[Dict[str, str]]] = {}
    for catalog_id, entries in by_catalog.items():
        entries.sort(key=lambda pair: pair[0])
        ordered[catalog_id] = [option for _, option in entries]
    return ordered


def _order(value: Any) -> int:
    try:
        return int(float(_text(value)))
    except (TypeError, ValueError):
        return 10_000


def _questions_by_variable(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    """The wording for each variable, keyed by variable id.

    Several rows may exist for one variable — one per language, or per country.
    All are kept: which one becomes the label is decided later, by the language
    the form is being imported in.
    """
    by_variable: Dict[str, List[Dict[str, str]]] = {}
    for row in rows:
        variable_id = _text(row.get("Variable ID"))
        text = _text(row.get("Question Text"))
        if not variable_id or not text:
            continue
        by_variable.setdefault(variable_id, []).append({
            "language": _text(row.get("Language Tag")).lower() or "",
            "text": text,
            "catalog_id": _text(row.get("Response Catalog ID")),
            "instructions": _text(row.get("Instructions / Enumerator Note")),
            "skip_logic": _text(row.get("Skip Logic Ref.")),
            "recall_period": _text(row.get("Recall Period")),
        })
    return by_variable


def _labels_by_target(rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str], Dict[str, str]]:
    """Translations the client already wrote, keyed by what they describe.

    These are carried into the form as they stand. Nothing is generated to fill
    a gap: a language the workbook does not translate simply is not translated.
    """
    by_target: Dict[Tuple[str, str], Dict[str, str]] = {}
    for row in rows:
        target_type = _text(row.get("Target Type")).lower()
        target_id = _text(row.get("Target ID"))
        language = _text(row.get("Language Tag")).lower()
        label = _text(row.get("Label"))
        if not target_id or not language or not label:
            continue
        if _text(row.get("Label Role")).lower() not in ("", "preferred"):
            continue
        by_target.setdefault((target_type, target_id), {})[language] = label
    return by_target


def _mappings_by_target(rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str], List[Dict[str, str]]]:
    """Standards the client has already mapped, kept as provenance."""
    by_target: Dict[Tuple[str, str], List[Dict[str, str]]] = {}
    for row in rows:
        target_id = _text(row.get("Target ID"))
        if not target_id:
            continue
        by_target.setdefault(
            (_text(row.get("Target Type")).lower(), target_id), []
        ).append({
            "vocabulary": _text(row.get("External Vocabulary")),
            "uri_or_code": _text(row.get("External URI / Code")),
            "label": _text(row.get("External Label")),
            "relation": _text(row.get("Mapping Relation")),
        })
    return by_target


def _field_type(data_type: str, has_options: bool) -> str:
    """The field type to store this variable in.

    A variable with a catalog is a choice whatever its declared type says, and a
    type the workbook uses that this reader does not recognise becomes text
    rather than a guess.
    """
    declared = DATA_TYPES.get(_text(data_type).lower())
    if has_options:
        return declared if declared in ("select", "multiselect") else "select"
    return declared or "text"


def _dominant_language(questions: Dict[str, List[Dict[str, str]]]) -> str:
    """The language the workbook's questions are written in.

    Counted from the question rows rather than assumed: a Spanish workbook must
    produce a Spanish form, and nothing here may change that.
    """
    counts: Dict[str, int] = {}
    for rows in questions.values():
        for row in rows:
            language = row["language"]
            if language:
                counts[language] = counts.get(language, 0) + 1
    if not counts:
        return ""
    return max(counts.items(), key=lambda pair: (pair[1], pair[0]))[0]


# --------------------------------------------------------------------------- #
# turning it into forms
# --------------------------------------------------------------------------- #
def build_forms(sheets: Dict[str, List[Dict[str, Any]]], source: str = "") -> List[Dict[str, Any]]:
    """One form per profile in the workbook.

    A profile names the variables it includes, their order and whether each is
    required, which is exactly a form. A workbook with no profiles still yields
    one form holding every variable, because a workbook of variables is still a
    definition of something.
    """
    variables = {
        _text(row.get("Variable ID")): row
        for row in sheets.get(VARIABLES, [])
        if _text(row.get("Variable ID"))
    }
    catalogs = _catalog_options(sheets.get(CATALOG_VALUES, []))
    catalog_names = {
        _text(row.get("Catalog ID")): _text(row.get("Catalog Name"))
        for row in sheets.get(CATALOGS, [])
    }
    units = {
        _text(row.get("Unit ID")): _text(row.get("Symbol")) or _text(row.get("Preferred Name"))
        for row in sheets.get(UNITS, [])
    }
    questions = _questions_by_variable(sheets.get(QUESTIONS, []))
    labels = _labels_by_target(sheets.get(LABELS, []))
    mappings = _mappings_by_target(sheets.get(MAPPINGS, []))

    language = _dominant_language(questions)

    memberships: Dict[str, Dict[str, Any]] = {}
    for row in sheets.get(PROFILES, []):
        profile_id = _text(row.get("Profile ID"))
        variable_id = _text(row.get("Variable ID"))
        if not profile_id or not variable_id:
            continue
        profile = memberships.setdefault(profile_id, {
            "profile_id": profile_id,
            "name": _text(row.get("Profile Name")) or profile_id,
            "version": _text(row.get("Profile Version")),
            "scope": _text(row.get("Use Case / Scope")),
            "members": [],
        })
        profile["members"].append({
            "variable_id": variable_id,
            "required": _text(row.get("Requirement Level")).lower() in REQUIRED_LEVELS,
            "order": _order(row.get("Display Order")),
        })

    if not memberships:
        memberships = {"_all": {
            "profile_id": "",
            "name": source or "Imported form",
            "version": "",
            "scope": "",
            "members": [
                {"variable_id": vid, "required": False, "order": index}
                for index, vid in enumerate(variables, start=1)
            ],
        }}

    forms = []
    for profile in memberships.values():
        forms.append(_build_one(
            profile, variables, catalogs, catalog_names, units,
            questions, labels, mappings, language, source,
        ))
    return forms


def _build_one(profile, variables, catalogs, catalog_names, units,
               questions, labels, mappings, language, source) -> Dict[str, Any]:
    fields = []
    taken = set()

    for member in sorted(profile["members"], key=lambda m: m["order"]):
        variable = variables.get(member["variable_id"])
        if not variable:
            # A profile naming a variable the workbook does not define is the
            # client's inconsistency to resolve, not this reader's to paper over.
            logger.warning("Profile %s names unknown variable %s",
                           profile["profile_id"], member["variable_id"])
            continue

        field = _build_field(
            variable, member, catalogs, catalog_names, units,
            questions, labels, mappings, taken,
        )
        if field:
            field["order"] = len(fields) + 1
            fields.append(field)

    definition: Dict[str, Any] = {
        "title": profile["name"],
        "description": profile["scope"],
        "fields": fields,
        "sections": [],
        # Where this came from, kept with the form so an imported definition can
        # always be traced back to the workbook and the profile it describes.
        "import_source": {
            "kind": "cimmyt_workbook",
            "file": source,
            "profile_id": profile["profile_id"],
            "profile_version": profile["version"],
        },
    }

    if language:
        # The workbook's language, carried through untouched. Nothing in this
        # reader translates, and nothing downstream may either.
        definition["default_language"] = language
        definition["languages"] = [language]

    translations = _form_translations(fields, labels, language)
    if translations:
        definition["translations"] = translations
        definition["languages"] = sorted(
            {language} | set(translations) if language else set(translations)
        )

    return definition


def _build_field(variable, member, catalogs, catalog_names, units,
                 questions, labels, mappings, taken) -> Optional[Dict[str, Any]]:
    variable_id = member["variable_id"]
    name_source = _text(variable.get("Preferred Variable Name")) or variable_id

    name = _slug(name_source) or _slug(variable_id)
    while name in taken:
        name = f"{name}_2"
    taken.add(name)

    asked = questions.get(variable_id) or []
    question = asked[0] if asked else None

    # The catalog the question answers with wins over the variable's own, since
    # the question is what a person is actually filling in.
    catalog_id = ""
    if question and question["catalog_id"]:
        catalog_id = question["catalog_id"]
    if not catalog_id:
        catalog_id = _text(variable.get("Catalog ID"))

    options = [dict(option) for option in catalogs.get(catalog_id, [])]

    field: Dict[str, Any] = {
        "name": name,
        # The wording the client wrote, in the language they wrote it in.
        "label": (question["text"] if question else "") or name_source,
        "type": _field_type(variable.get("Data Type"), bool(options)),
        "required": member["required"],
        "help_text": (question or {}).get("instructions") or _text(
            variable.get("Operational Definition")
        ),
        "options": options,
        "validation": {},
    }

    source_meta: Dict[str, Any] = {"variable_id": variable_id}
    for key, column in (
        ("concept_id", "Concept ID"),
        ("variable_name", "Preferred Variable Name"),
        ("data_type", "Data Type"),
        ("observation_entity", "Observation Entity"),
        ("measurement_role", "Measurement Role"),
        ("temporal_basis", "Temporal Basis"),
        ("method", "Collection / Calculation Method"),
        ("status", "Status"),
        ("version", "Version"),
    ):
        value = _text(variable.get(column))
        if value:
            source_meta[key] = value

    unit_id = _text(variable.get("Unit ID"))
    if unit_id:
        source_meta["unit_id"] = unit_id
        source_meta["unit"] = units.get(unit_id, "")

    if catalog_id:
        # The client's catalog is the authority for this field's values. Recorded
        # so nothing downstream replaces the list with a standard's own.
        source_meta["catalog_id"] = catalog_id
        source_meta["catalog_name"] = catalog_names.get(catalog_id, "")
        source_meta["catalog_is_client_controlled"] = True

    if question:
        if question["language"]:
            source_meta["language"] = question["language"]
        if question["skip_logic"]:
            # Kept exactly as written. This reader records the client's condition;
            # it does not interpret or re-express it.
            source_meta["skip_logic"] = question["skip_logic"]
        if question["recall_period"]:
            source_meta["recall_period"] = question["recall_period"]

    existing = mappings.get(("variable", variable_id)) or mappings.get(
        ("concept", _text(variable.get("Concept ID")))
    )
    if existing:
        source_meta["external_mappings"] = existing

    field["source"] = source_meta
    return field


def _form_translations(fields, labels, language) -> Dict[str, Dict[str, Any]]:
    """The client's own translations, arranged the way a form stores them.

    Only what the workbook already contains. A label the client has not
    translated stays untranslated — this import never generates one.
    """
    translations: Dict[str, Dict[str, Any]] = {}

    for field in fields:
        source = field.get("source") or {}
        for target in (("variable", source.get("variable_id")),
                       ("concept", source.get("concept_id"))):
            if not target[1]:
                continue
            for code, label in (labels.get(target) or {}).items():
                if not code or code == language:
                    continue
                block = translations.setdefault(code, {"fields": {}})
                block["fields"].setdefault(field["name"], {"label": label})

    return translations


def import_workbook(data: bytes, source: str = "") -> List[Dict[str, Any]]:
    """The whole read: bytes in, one form definition per profile out."""
    return build_forms(read_workbook(data), source=source)
