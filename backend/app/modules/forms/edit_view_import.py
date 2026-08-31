"""
Import a client "Edit view" Excel workbook into Standard Form JSON.

This reader is intentionally separate from excel_import.py.

excel_import.py handles the CIMMYT Controlled Vocabulary workbook
(03_Variables, 04_Value_Catalogs, 05_Catalog_Values, 14_Profiles, etc.).

This reader handles the client's form-oriented "Edit view" workbook.

Important:
- The workbook is authoritative for fields and labels.
- No LLM is used.
- No values are invented.
- Existing client catalogs are preserved.
- Existing standards enrichment happens later in standard_forms.py.
"""

from __future__ import annotations

import io
import re
import warnings
from typing import Any, Dict, List, Optional

import openpyxl


class EditViewWorkbookProblem(ValueError):
    """The workbook is not a supported Edit View workbook."""


def _text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip()

    if text.lower() in ("nan", "none", "-"):
        return ""

    return text


def _slug(value: str) -> str:
    """
    Create a stable form field name.

    We intentionally keep this simple because the source workbook's
    VARIABLE column is the authority.
    """
    value = _text(value).lower()

    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value)
    value = value.strip("_")

    return value


def _bool(value: Any) -> bool:
    text = _text(value).lower()

    return text in {
        "yes",
        "y",
        "true",
        "1",
        "required",
        "mandatory",
        "si",
        "sí",
    }


def _find_header_row(ws) -> Optional[int]:
    """
    Find the actual header row.

    The Edit View workbook can have title/instruction rows before the
    actual table header.
    """

    required_markers = {
        "VARIABLE",
        "FIELD TYPE",
        "REQUIRED",
    }

    for row_index, row in enumerate(
        ws.iter_rows(values_only=True),
        start=1,
    ):

        values = {
            _text(value).upper()
            for value in row
            if _text(value)
        }

        if len(required_markers.intersection(values)) >= 2:
            return row_index

    return None


def _read_rows(ws) -> List[Dict[str, Any]]:
    header_row = _find_header_row(ws)

    if header_row is None:
        return []

    rows = list(
        ws.iter_rows(
            min_row=header_row,
            values_only=True,
        )
    )

    if not rows:
        return []

    headers = [_text(value) for value in rows[0]]

    result: List[Dict[str, Any]] = []

    for row in rows[1:]:

        record: Dict[str, Any] = {}

        # A condition is written across several cells under one heading:
        #
        #   LOGIC     |        |    |
        #   SHOW IF   | tipo_c | IS | Persona_fisica
        #
        # Only the first cell has a header, so the columns after it belong to
        # the same heading. Dropping them would keep "SHOW IF" and lose the
        # condition itself, which is the part worth preserving.
        last_header = ""

        for header, value in zip(headers, row):

            value = _text(value)

            if header:
                last_header = header
                record[header] = value
                continue

            if last_header and value:
                record[last_header] = f"{record.get(last_header, '')} {value}".strip()

        if any(record.values()):
            result.append(record)

    return result


def _find_edit_view_sheet(workbook):
    """
    Find a sheet that looks like the client's Edit View form.

    Prefer a sheet literally named "Edit view".
    Otherwise detect one by its headers.
    """

    for name in workbook.sheetnames:

        if name.strip().lower() == "edit view":
            return workbook[name]

    for name in workbook.sheetnames:

        ws = workbook[name]

        if _find_header_row(ws) is not None:
            return ws

    return None


def is_edit_view_workbook(data: bytes) -> bool:
    """
    Return True when the workbook contains an Edit View-style sheet.
    """

    try:

        with warnings.catch_warnings():

            warnings.simplefilter("ignore")

            workbook = openpyxl.load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=True,
            )

            ws = _find_edit_view_sheet(workbook)

            if ws is None:
                return False

            header_row = _find_header_row(ws)

            if header_row is None:
                return False

            # Read only the header.
            row = next(
                ws.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    values_only=True,
                ),
                (),
            )

            headers = {
                _text(value).upper()
                for value in row
                if _text(value)
            }

            return (
                "VARIABLE" in headers
                and "FIELD TYPE" in headers
            )

    except Exception:
        return False


def read_workbook(
    data: bytes,
    source: str = "",
) -> List[Dict[str, Any]]:
    """
    Read an Edit View workbook and return Standard Form definitions.
    """

    try:

        with warnings.catch_warnings():

            warnings.simplefilter("ignore")

            workbook = openpyxl.load_workbook(
                io.BytesIO(data),
                read_only=True,
                data_only=True,
            )

            ws = _find_edit_view_sheet(workbook)

            if ws is None:
                raise EditViewWorkbookProblem(
                    "No Edit View sheet was found. "
                    "Expected a sheet containing VARIABLE and FIELD TYPE columns."
                )

            rows = _read_rows(ws)

    except EditViewWorkbookProblem:
        raise

    except Exception as exc:
        raise EditViewWorkbookProblem(
            f"Could not read the Edit View workbook: {exc}"
        ) from exc

    if not rows:
        raise EditViewWorkbookProblem(
            "The Edit View sheet contains no form fields."
        )

    return build_form(rows, source=source)


# How the client writes a type, folded onto ours.
#
# Keys are compared after folding: lower case, with spaces, hyphens and
# underscores removed. So "select1", "Select 1" and "select_1" are one key.
TYPE_MAPPING = {
    "text": "text",
    "string": "text",
    "html": "text",
    "memo": "textarea",
    "textarea": "textarea",
    "longtext": "textarea",

    "integer": "number",
    "int": "number",
    "number": "number",
    "numeric": "decimal",
    "decimal": "decimal",
    "float": "decimal",

    "date": "date",
    "datetime": "datetime",
    "time": "time",

    "boolean": "boolean",
    "bool": "boolean",

    # One answer from a controlled list. "select1" is the ODK/XLSForm spelling
    # the client's Edit view uses, and it is the common case in their workbooks.
    "select": "select",
    "select1": "select",
    "selectone": "select",
    "singleselect": "select",
    "dropdown": "select",
    "choice": "select",
    "categorical": "select",
    "code": "select",

    "multiselect": "multiselect",
    "selectmultiple": "multiselect",
    "selectn": "multiselect",
    "checkbox": "multiselect",
}

# The types above that mean "pick from a controlled list". A field of one of
# these is answered from a catalog, never from free text.
CONTROLLED_TYPES = {"select", "multiselect"}


def _fold_type(source_type: str) -> str:
    """Fold a written type so "select1", "Select 1" and "select_1" agree."""
    return re.sub(r"[\s_\-]+", "", _text(source_type).lower())


def _field_type(source_type: str, catalog: str) -> str:
    """
    Convert the client's field type into our Form Builder type.

    We only make straightforward mappings.
    Unknown types become text rather than being guessed.
    """

    mapped = TYPE_MAPPING.get(_fold_type(source_type))

    if mapped:
        return mapped

    # A catalog means the field is controlled.
    if catalog:
        return "select"

    return "text"


def _split_options(value: str) -> List[Dict[str, str]]:
    """
    Parse simple inline option lists if the client explicitly put
    them into the workbook.

    We DO NOT invent values.

    Supported examples:

        Yes | No | Unknown
        Yes, No, Unknown
        Y:Yes | N:No | UNK:Unknown
    """

    value = _text(value)

    if not value:
        return []

    parts = re.split(r"\s*[|;,]\s*", value)

    options: List[Dict[str, str]] = []

    for part in parts:

        part = _text(part)

        if not part:
            continue

        if ":" in part:

            code, label = part.split(":", 1)

            code = _text(code)
            label = _text(label)

            if code and label:
                options.append(
                    {
                        "label": label,
                        "value": code,
                    }
                )

        else:

            options.append(
                {
                    "label": part,
                    "value": _slug(part),
                }
            )

    return options


def _first_value(
    row: Dict[str, Any],
    *names: str,
) -> str:

    upper = {
        _text(key).upper(): _text(value)
        for key, value in row.items()
    }

    for name in names:

        value = upper.get(name.upper())

        if value:
            return value

    return ""


# The conditions a client's LOGIC column is written in, and what each means.
#
#     SHOW IF rcl_tipo_colaborador_c IS Persona_fisica
#     HIDE IF estado_c IS NOT Activo
#
# Only these shapes are read. Anything else is kept verbatim in
# `source.skip_logic`, exactly as before, and reported — guessing at a condition
# would change what the client wrote, which is worse than leaving it unenforced.
SKIP_LOGIC = re.compile(
    r"^\s*(?P<action>SHOW|HIDE)\s+IF\s+"
    r"(?P<field>[^\s]+)\s+"
    r"(?P<operator>IS\s+NOT|IS|!=|<>|=|==)\s+"
    r"(?P<value>.+?)\s*$",
    re.IGNORECASE,
)

SKIP_OPERATORS = {
    "is": "equals",
    "=": "equals",
    "==": "equals",
    "is not": "not_equals",
    "!=": "not_equals",
    "<>": "not_equals",
}


def _rule_from_logic(logic: str, field_name: str) -> Optional[Dict[str, Any]]:
    """The client's condition as a rule this application can act on.

    None when it is not one of the shapes above. The original text is kept
    either way; this only adds an enforceable reading of it where one is certain.
    """
    match = SKIP_LOGIC.match(_text(logic))

    if not match:
        return None

    operator = SKIP_OPERATORS.get(
        re.sub(r"\s+", " ", match.group("operator").strip().lower())
    )

    if not operator:
        return None

    return {
        "conditions": [{
            # The client's own variable name, slugified the same way the field
            # it refers to was — so the rule names a question this form has.
            "field": _slug(match.group("field")),
            "operator": operator,
            "value": _text(match.group("value")),
        }],
        "logic": "AND",
        "action": match.group("action").lower(),
        "target": {"type": "field", "name": field_name},
    }


def _make_section_key(value: str) -> str:
    return _slug(value) or "default_section"


# The columns a label can be written in, and the language each one is written
# in. The workbook's own headings are the evidence — "LABEL SPAN" beside
# "ETIQUETA ENG" says the form is Spanish and carries an English translation.
LABEL_COLUMNS = [
    ("LABEL SPAN", "es"),
    ("ETIQUETA SPAN", "es"),
    ("ETIQUETA", "es"),
    ("LABEL ESP", "es"),
    ("LABEL ES", "es"),
    ("LABEL FR", "fr"),
    ("ETIQUETTE", "fr"),
    ("LABEL PT", "pt"),
    ("ROTULO", "pt"),
    ("LABEL", "en"),
    ("LABEL ENG", "en"),
    ("ENGLISH LABEL", "en"),
    ("QUESTION", "en"),
    ("QUESTION TEXT", "en"),
]

ENGLISH_LABEL_COLUMNS = (
    "ETIQUETA ENG",
    "LABEL ENG",
    "ENGLISH LABEL",
    "LABEL EN",
)

# What a LANGUAGE column may say, folded onto a language tag.
LANGUAGE_NAMES = {
    "english": "en",
    "spanish": "es",
    "espanol": "es",
    "español": "es",
    "castellano": "es",
    "french": "fr",
    "francais": "fr",
    "français": "fr",
    "portuguese": "pt",
    "portugues": "pt",
    "português": "pt",
}


def _language_tag(value: str) -> str:
    """Read a written language as a tag. "Spanish", "es-MX" and "ES" are all es."""
    value = _text(value).lower()

    if not value:
        return ""

    if value in LANGUAGE_NAMES:
        return LANGUAGE_NAMES[value]

    # es-MX, pt_BR: the region does not change which words are on the form.
    tag = re.split(r"[-_]", value)[0]

    return tag if len(tag) == 2 else ""


def _headers_of(rows: List[Dict[str, Any]]) -> set:
    headers = set()

    for row in rows:
        headers.update(_text(key).upper() for key in row)

    return headers


def _default_language(rows: List[Dict[str, Any]]) -> str:
    """
    Which language the workbook is written in.

    Never translated and never assumed: a LANGUAGE column decides it if the
    workbook has one, otherwise the heading the labels sit under does. English
    is the answer only when the workbook actually reads as English — a Spanish
    workbook must import as a Spanish form.
    """

    for row in rows:

        tag = _language_tag(
            _first_value(
                row,
                "LANGUAGE",
                "LANGUAGE TAG",
                "LANG",
                "IDIOMA",
            )
        )

        if tag:
            return tag

    headers = _headers_of(rows)

    for column, language in LABEL_COLUMNS:

        if column in headers:
            return language

    return "en"


def build_form(
    rows: List[Dict[str, Any]],
    source: str = "",
) -> List[Dict[str, Any]]:
    """
    Convert Edit View rows into one Standard Form definition.
    """

    fields: List[Dict[str, Any]] = []

    sections: List[Dict[str, Any]] = []
    section_seen = set()

    used_names = set()

    default_language = _default_language(rows)

    # The client's own English labels, when the form itself is in another
    # language. Carried across as a translation they already wrote — nothing
    # here translates anything.
    english_labels: Dict[str, Dict[str, str]] = {}

    # Conditions read from the LOGIC column, and the ones that could not be.
    rules: List[Dict[str, Any]] = []
    unread_logic: List[str] = []

    for index, row in enumerate(rows, start=1):

        variable = _first_value(
            row,
            "VARIABLE",
            "VARIABLE NAME",
            "FIELD",
            "FIELD NAME",
        )

        if not variable:
            continue

        name = _slug(variable)

        if not name:
            continue

        # Prevent duplicate field names.
        original_name = name
        suffix = 2

        while name in used_names:

            name = f"{original_name}_{suffix}"
            suffix += 1

        used_names.add(name)

        # Same order as LABEL_COLUMNS, so the label read is the one whose
        # heading decided the form's language.
        label = _first_value(row, *[column for column, _ in LABEL_COLUMNS])

        english_label = _first_value(row, *ENGLISH_LABEL_COLUMNS)

        field_type_source = _first_value(
            row,
            "FIELD TYPE",
            "TYPE",
            "DATA TYPE",
        )

        required_source = _first_value(
            row,
            "REQUIRED",
            "REQUIRED?",
            "REQUIRED FIELD",
        )

        catalog_id = _first_value(
            row,
            "CATALOG",
            "CATALOG ID",
            "RESPONSE CATALOG",
            "RESPONSE CATALOG ID",
        )

        father_list = _first_value(
            row,
            "FATHER LIST",
            "PARENT LIST",
            "PARENT CODE",
            "PARENT",
        )

        panel = _first_value(
            row,
            "PANEL",
            "SECTION",
            "GROUP",
        )

        logic = _first_value(
            row,
            "LOGIC",
            "SKIP LOGIC",
            "SKIP LOGIC REF.",
        )

        location = _first_value(
            row,
            "LOCATION",
        )

        size_annotation = _first_value(
            row,
            "SIZE/ANOTATION",
            "SIZE",
            "ANNOTATION",
        )

        apply_to = _first_value(
            row,
            "APPLY TO",
        )

        options_text = _first_value(
            row,
            "OPTIONS",
            "VALUES",
            "ALLOWED VALUES",
        )

        options = _split_options(options_text)

        field_type = _field_type(
            field_type_source,
            catalog_id,
        )

        # A client catalog is dynamic and authoritative.
        if catalog_id:

            # If the client has explicitly said this is a catalog,
            # don't let standards replace its choices.
            field_type = "select"

        field: Dict[str, Any] = {
            "name": name,
            "label": label or english_label or variable,
            "type": field_type,
            "required": _bool(required_source),
            "placeholder": "",
            "help_text": _first_value(row, "HELP TEXT", "HELP", "AYUDA"),
            "default": None,
            "section": _make_section_key(panel),
            "options": options,
            "validation": {},
            "order": index,
        }

        source_meta: Dict[str, Any] = {
            "kind": "edit_view_workbook",
            "source_variable": variable,
        }

        if field_type_source:
            source_meta["field_type"] = field_type_source

        if location:
            source_meta["location"] = location

        if size_annotation:
            source_meta["size_annotation"] = size_annotation

        if apply_to:
            source_meta["apply_to"] = apply_to

        if father_list:
            source_meta["father_list"] = father_list

        if logic:
            # Kept verbatim, as it always was. A reading of it is added beside
            # it when the shape is one we are certain of.
            source_meta["skip_logic"] = logic

            rule = _rule_from_logic(logic, name)
            if rule:
                rules.append(rule)
            else:
                unread_logic.append(f"{variable}: {logic}")

        # The client said this answer comes from a controlled list. Recorded
        # even when we cannot see the list, so the field stays a dropdown
        # instead of quietly becoming a text box that accepts anything.
        if field_type in CONTROLLED_TYPES:
            source_meta["controlled_list"] = True

        # ---------------------------------------------------------------
        # Client-controlled catalog
        # ---------------------------------------------------------------

        if catalog_id:

            # The catalog is named, not copied. Its values are read from
            # PostgreSQL when the form is drawn, so the client's own list stays
            # the one answer of record — no standard and no model supplies it.
            field["options_from"] = {
                "source": "client_catalog",
                "catalog": catalog_id,
            }

            # Nothing on the form carries the values, so anything written into
            # OPTIONS would compete with the catalog. The catalog wins.
            field["options"] = []

            source_meta["catalog_id"] = catalog_id
            source_meta["catalog_is_client_controlled"] = True

            # ------------------------------------------------------------
            # Parent dependency
            #
            # A municipality list belongs to a state. `depends_on` names the
            # field holding that state, and the catalog is read again for the
            # answer given there. Only meaningful alongside a catalog: there is
            # nothing to narrow without one.
            # ------------------------------------------------------------

            if father_list:
                field["options_from"]["depends_on"] = _slug(father_list)

        field["source"] = source_meta

        # The client's own English wording, kept as a translation beside their
        # label rather than in place of it.
        if (
            english_label
            and english_label != field["label"]
            and default_language != "en"
        ):
            english_labels[name] = {"label": english_label}

        # ---------------------------------------------------------------
        # Sections
        # ---------------------------------------------------------------

        section_key = _make_section_key(panel)

        if section_key not in section_seen:

            section_seen.add(section_key)

            sections.append(
                {
                    "key": section_key,
                    "title": panel or "General",
                    "description": "",
                }
            )

        fields.append(field)

    if not fields:
        raise EditViewWorkbookProblem(
            "The Edit View workbook contains no usable VARIABLE rows."
        )

    title = source.rsplit("/", 1)[-1]
    title = title.rsplit("\\", 1)[-1]

    if "." in title:
        title = title.rsplit(".", 1)[0]

    languages = [default_language]
    translations: Dict[str, Any] = {}

    if english_labels:
        languages.append("en")
        translations["en"] = {"fields": english_labels}

    # A rule naming a question this workbook did not define cannot be acted on.
    # The text stays on the field either way.
    known = {f["name"] for f in fields}
    for rule in list(rules):
        if rule["conditions"][0]["field"] not in known:
            unread_logic.append(
                f"{rule['target']['name']}: refers to '{rule['conditions'][0]['field']}', "
                f"which is not a question in this workbook"
            )
            rules.remove(rule)

    definition: Dict[str, Any] = {
        "title": title or "Imported Standard Form",
        "description": "",
        "fields": fields,
        "sections": sections,
        "rules": rules,
        "languages": languages,
        "default_language": default_language,
        "translations": translations,
        "import_source": {
            "kind": "edit_view_workbook",
            "file": source,
            # Conditions written in a shape this reader does not recognise. They
            # are still on their fields as the client wrote them; they are simply
            # not enforced, and saying so is better than appearing to enforce
            # something we guessed at.
            **({"unread_logic": unread_logic} if unread_logic else {}),
        },
    }

    return [definition]